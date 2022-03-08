# download_playlist.py
import argparse
from pathlib import Path
ROOT = Path(__file__).parent.absolute()

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter 
import youtube_dl


# argparse code for passing in the playlist to download with -p or --playlist
parser = argparse.ArgumentParser(description='Use this file to download and convert the videos in a Youtube playlist to .mp3 files.')
parser.add_argument('-p', '--playlist', required=True, help='the link to the Youtube playlist that will be downloaded')
args = parser.parse_args()

# FLAG is a boolean flag used for the hooks that print the code's current status.
FLAG = False


# I'll be honest, I have zero clue what the MyLogger class is for.
class MyLogger(object):
	def debug(self, msg):
		pass

	def warning(self, msg):
		pass

	def error(self, msg):
		print(msg)


# Main method that drives the code
def main():

	# ydl is the YoutubeDL object that is used to fetch data from Youtube.
	ydl = youtube_dl.YoutubeDL(
	{
		'format': 'bestaudio/best',
		'postprocessors': [{
			'key': 'FFmpegExtractAudio',
			'preferredcodec': 'mp3',
		}],
		'logger': MyLogger(),
		'progress_hooks': [my_hook],
		'outtmpl': str(ROOT / 'tracks' / '%(title)s.%(ext)s'),
	})

	# botched_vids[] is an array that will store any songs that encounter a problem while downloading.
	botched_vids = []

	with ydl:

		# videos is an array that holds all the videos in the Youtube Playlist
		videos = ydl.extract_info(args.playlist, download=False)['entries']

		# For every video, try to download it via the webpage_url value stored in video
		for video in videos:
			try:
				ydl.download([video['webpage_url']])

			# In case the user tries to use CTRL-C to close out the program, this try-catch would prevent
			#	 it from triggering, so this special except clause is added.
			except KeyboardInterrupt:
				print('\033[91mYou triggered a KeyBoardInterrupt exception.\033[0m')
				print('\033[91mShutting down download_playlist.py\033[0m')
				return

			# If something goes wrong, print the following message and append the video to botched_vids[]
			#	 for review later.
			except:
				print('\033[93m  Skipping this song since something went wrong...\033[0m')
				botched_vids.append(video['title'])

	# If there were any videos that couldn't be downloaded, this is where they're be displayed.
	if botched_vids:
		print("The following songs couldn't be downloaded:")
		for title in botched_vids:
			print(f'  {title}')

	# Finally, create the Excel file that will store the song information.
	# This information will need to be manually insterted by the user later.
	wb = Workbook()
	ws = wb.active
	ws.title = 'Song Information'
	headers = ["", "Franchise", "Game", "Song"]
	for i in range(1, 4):
		ws.column_dimensions[get_column_letter(i)].width = 33
		ws.cell(1, i).font = Font(bold=True, underline='single')
		ws.cell(1, i).value = headers[i]

	wb.save(ROOT / 'tracks' / 'song_info.xlsx')


# Method for showing progess of download/conversion
def my_hook(d):
	global FLAG

	if not FLAG and d['status'] == 'downloading':
		FLAG = not FLAG
		print(f"\n\033[92m  Downloading {d['filename'].split(chr(92))[-1].split('.')[0]}...\033[0m")
	if d['status'] == 'finished':
		FLAG = False
		print('  Converting to .mp3...')


if __name__ == '__main__':
	print('\nRunning download_playlist.py')
	main()